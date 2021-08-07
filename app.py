import shutil
from openpyxl import load_workbook
import os
import glob
from datetime import datetime
import PySimpleGUI as sg
import json
import codecs
import webbrowser

# if app don't work
# pip install openpyxl


def app(data):
    searchFolder = data["searchFolder"]
    destinationPath = data["destinationPath"]
    xlsName = data["xlsName"]
    preserveOriginalFilename = data["preserveOriginalFilename"]

    searchSubfolders = "\**"
    origin = searchFolder + searchSubfolders
    destination = os.path.join(destinationPath, nowCurrentTime())

    xls = readXlsAndReturnValues(xlsName)
    manipulateXls(xls, destination, origin, preserveOriginalFilename)


def manipulateXls(xls, destination, origin, preserveOriginalFilename):
    for item in xls:
        folderName = item[0]
        if folderName != None:
            keywordsTemp = item[1:]
            destinationTemp = os.path.join(destination, folderName)
            copyAllFiles(
                origin,
                destinationTemp,
                keywordsTemp,
                folderName,
                preserveOriginalFilename,
            )
        else:
            print("\n\n\n!!--!!--!!\t\t Empty row A: Folder Name\n\n")


def nowCurrentTime():
    nowTime = datetime.now()
    nowTime = nowTime.strftime("%Y-%m-%d__%H-%M-%S")
    return str(nowTime)


def readXlsAndReturnValues(xlsName):
    newFilteredList = []
    wb = load_workbook(xlsName)
    sheet = wb.active
    rows_iter = sheet.iter_rows(max_col=sheet.max_column, max_row=sheet.max_row)
    allValuesFromXLS = [[cell.value for cell in list(row)] for row in rows_iter]

    # remove "None" from values
    for index in range(len(allValuesFromXLS)):
        tempList = list(filter(None, allValuesFromXLS[index]))
        if len(tempList) > 0:
            tempList = list(dict.fromkeys(tempList))  # remove duplicates
            newFilteredList.append(tempList)
    return newFilteredList


def createFolderIfNotExist(pathToNewFolder):
    if not os.path.exists(pathToNewFolder):
        os.makedirs(pathToNewFolder)
        print("\n++ Folders created! \n\t", pathToNewFolder, "\n")
    else:
        print("\n-- Folder already exist,  \n\t", pathToNewFolder, "\n")


def filterArray(array, keyWords, path):
    # array and keyWords will be reduced to lowercase
    array = [each_string.lower() for each_string in array]
    keyWords = [each_keyWords.lower() for each_keyWords in keyWords]
    for key in keyWords:
        array = [item for item in array if key in item]
    if len(array) == 0 or not os.path.isfile(path):
        return False
    else:
        return True


def copyAllFiles(origin, destination, keyWords, folderName, preserveOriginalFilename):
    listOfAllFilesFullPath = []
    keyWords = [x for x in keyWords if x is not None]
    print("\n\n------ FOLDER NAME:\t\t", folderName)
    print("------ KEYWORDS:\t", keyWords)
    if not keyWords:
        return print("!!--!!--!! \t\t\t Keywords not found ")
    print("\n++ Files found: \t")

    for f in glob.glob(origin, recursive=True):
        if filterArray([os.path.basename(f)], keyWords, f):
            listOfAllFilesFullPath.append(f)
            print("\t" + os.path.basename(f))
    if len(listOfAllFilesFullPath) == 0:
        print("\t-- !! -- \t Nothing -> List is empty\n\n\n")
        return False

    try:
        createFolderIfNotExist(destination)
        id = 0
        # copy and rename
        for fileName in listOfAllFilesFullPath:
            if os.walk(fileName):
                shutil.copy(fileName, destination)
                fileTempPath = os.path.join(destination, os.path.basename(fileName))
                keyWordBuilder = "_".join(keyWords)
                extension = os.path.splitext(fileName)[1]
                if not (preserveOriginalFilename):
                    if id == 0:
                        os.rename(
                            fileTempPath, destination + "/" + keyWordBuilder + extension
                        )
                    else:
                        os.rename(
                            fileTempPath,
                            destination
                            + "/"
                            + keyWordBuilder
                            + "__("
                            + str(id)
                            + ")"
                            + extension,
                        )
                else:
                    if id == 0:
                        os.rename(
                            fileTempPath,
                            destination
                            + "/"
                            + str(os.path.basename(fileName))
                            + "___#"
                            + keyWordBuilder
                            + "#___"
                            + extension,
                        )
                    else:
                        os.rename(
                            fileTempPath,
                            destination
                            + "/"
                            + str(os.path.basename(fileName))
                            + "___#"
                            + keyWordBuilder
                            + "#__("
                            + str(id)
                            + ")"
                            + extension,
                        )

                id = id + 1
        print("++ Files copied successfully \n\n\n")
        id = 0
    except:
        print("\n--!!!--\t Fail during copying files")


def loadJson():
    with codecs.open("userData.json", "r", "utf-8") as jsonFile:
        data = json.load(jsonFile)
        return data


def writeJson(data):
    newDict = {
        "searchFolder": data["searchFolder"],
        "destinationPath": data["destinationPath"],
        "xlsName": data["xlsName"],
        "preserveOriginalFilename": data["preserveOriginalFilename"],
        "generateTimestamp": data["generateTimestamp"],
    }
    with codecs.open("userData.json", "w", "utf-8") as jsonFile:
        json.dump(newDict, jsonFile, ensure_ascii=False, indent=4)


def gui(data):
    sizeText = (25, 1)
    sizeInput = (100, 1)
    buttonsSize = (20, 1)
    sg.theme("DarkBlue3")
    layout = [
        [
            sg.Text("Search Folders (with subfolders)", size=sizeText),
            sg.InputText(
                size=sizeInput, default_text=data["searchFolder"], key="searchFolder"
            ),
            sg.FolderBrowse(),
            sg.Button("Open Search Folder", size=buttonsSize, button_color="#56a653"),
        ],
        [
            sg.Text("Destination Path (where copy files)", size=sizeText),
            sg.InputText(
                size=sizeInput,
                default_text=data["destinationPath"],
                key="destinationPath",
            ),
            sg.FolderBrowse(),
            sg.Button(
                "Open Destination Folder", size=buttonsSize, button_color="#56a653"
            ),
        ],
        [
            sg.Text("Xls Name (Excel file) ", size=sizeText),
            sg.InputText(size=sizeInput, default_text=data["xlsName"], key="xlsName"),
            sg.FileBrowse(
                file_types=(("Excel", ("*.xlsx", "*.xlsm", "*.xltx", "*.xltm")),)
            ),
            sg.Button("Open Excel File", size=buttonsSize, button_color="#56a653"),
        ],
        [
            sg.Checkbox(
                default=(data["preserveOriginalFilename"]),
                text="Preserve Original Filename (TODO CORRECTLY WORK)",
                key="preserveOriginalFilename",
            ),
        ],
        [
            sg.Checkbox(
                default=(data["generateTimestamp"]),
                text="Generate Timestamp (date and hour) (TODO: DONT WORK)",
                key="generateTimestamp",
            ),
        ],
        [
            sg.Button("Ok", size=buttonsSize, button_color="green"),
            sg.Button("Read XLS", size=buttonsSize, button_color="brown"),
            sg.Button("Close", size=buttonsSize, button_color="#541001"),
        ],
    ]

    window = sg.Window("Excel Finder", layout)
    while True:
        event, values = window.read()
        if event == "Close":  # if user closes window or clicks cancel
            writeJson(values)
            print("\nAPP CLOSED, DATA SAVED")
            break
        elif event == sg.WIN_CLOSED:
            print("\nAPP CLOSED, SETTINGS DISCARDED")
            break
        elif event == "Open Search Folder":
            path = values["searchFolder"]
            if os.path.isdir(path):
                webbrowser.open(os.path.realpath(path))
            else:
                print("Search: Path don't exist")
        elif event == "Open Destination Folder":
            path = values["destinationPath"]
            if os.path.isdir(path):
                webbrowser.open(os.path.realpath(path))
            else:
                print("Destination: Path don't exist")
        elif event == "Open Excel File":
            path = values["xlsName"]
            print(path)
            if os.path.isfile(path):
                # this should be async
                sg.popup("Save and Close Excel File to continue...")
                os.system(os.path.realpath(path))
            else:
                print("Excel: File don't exist")
        elif event == "Ok":
            writeJson(values)
            print("\n\nSCRIPT START")
            app(values)
        elif event == "Read XLS":
            guiReadXls(values)


def guiReadXls(data):
    xls = readXlsAndReturnValues(data["xlsName"])
    print("\n******************\n\n** Excel Values:")
    for index in range(len(xls)):
        print("\n\nFolder name:\t", xls[index][0])
        print("tags:")
        for item in range(1, len(xls[index])):
            print("\t", xls[index][item])
    print("\n** end of XLS")


# app()

data = loadJson()

gui(data)
